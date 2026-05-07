import csv
import io
import json
import os
from typing import Any, Iterable, List, Sequence

from python_rag.core.error_codes import ERR_CELERY_ERROR
from python_rag.core.errors import AppError


SUPPORTED_DOCUMENT_EXTENSIONS = (
    "md",
    "txt",
    "json",
    "csv",
    "pdf",
    "docx",
    "xlsx",
)

PLAIN_TEXT_DOCUMENT_EXTENSIONS = {"md", "txt"}


def get_document_extension(filename: str) -> str:
    _, ext = os.path.splitext(filename or "")
    return ext.lower().lstrip(".")


def supported_document_extensions_text() -> str:
    return ", ".join(f".{ext}" for ext in SUPPORTED_DOCUMENT_EXTENSIONS)


def is_supported_document_filename(filename: str) -> bool:
    return get_document_extension(filename) in SUPPORTED_DOCUMENT_EXTENSIONS


def validate_supported_document_filename(filename: str) -> None:
    if is_supported_document_filename(filename):
        return
    raise AppError(
        ERR_CELERY_ERROR,
        "unsupported document format; currently supported: %s"
        % supported_document_extensions_text(),
    )


def _read_binary_file(path: str) -> bytes:
    if not os.path.exists(path):
        raise AppError(ERR_CELERY_ERROR, "document file does not exist")

    with open(path, "rb") as f:
        return f.read()


def _decode_text_bytes(raw: bytes) -> str:
    if not raw:
        return ""

    for encoding in ("utf-8", "utf-8-sig", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except Exception:
            continue

    return raw.decode("utf-8", errors="ignore")


def _extract_text_from_plain_file(path: str) -> str:
    return _decode_text_bytes(_read_binary_file(path))


def _compact_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(line.strip() for line in text.split("\n") if line.strip()).strip()


def _escape_markdown_table_cell(value: Any) -> str:
    return _compact_cell_text(value).replace("|", "\\|")


def _normalize_table_rows(rows: Iterable[Sequence[Any]]) -> List[List[str]]:
    normalized_rows: List[List[str]] = []
    max_width = 0

    for row in rows:
        cells = [_escape_markdown_table_cell(cell) for cell in row]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        normalized_rows.append(cells)
        max_width = max(max_width, len(cells))

    if max_width <= 0:
        return []

    for row in normalized_rows:
        if len(row) < max_width:
            row.extend([""] * (max_width - len(row)))
    return normalized_rows


def _rows_to_markdown_table(rows: Iterable[Sequence[Any]], title: str | None = None) -> str:
    normalized_rows = _normalize_table_rows(rows)
    if not normalized_rows:
        return ""

    header = normalized_rows[0]
    header = [
        cell if cell else "Column {0}".format(index + 1)
        for index, cell in enumerate(header)
    ]
    body_rows = normalized_rows[1:]

    parts: List[str] = []
    if title:
        parts.append(title)
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join(["---"] * len(header)) + " |")
    for row in body_rows:
        parts.append("| " + " | ".join(row) + " |")

    return "\n".join(parts)


def _extract_text_from_csv(path: str) -> str:
    raw_text = _decode_text_bytes(_read_binary_file(path))
    if not raw_text.strip():
        return ""

    sample = raw_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel

    try:
        reader = csv.reader(io.StringIO(raw_text), dialect)
        table = _rows_to_markdown_table(reader, title="[CSV Table]")
    except Exception:
        table = ""
    return table or raw_text


def _extract_text_from_json(path: str) -> str:
    raw_text = _decode_text_bytes(_read_binary_file(path))
    if not raw_text.strip():
        return ""

    try:
        data = json.loads(raw_text)
    except Exception:
        return raw_text

    if isinstance(data, list) and data and all(isinstance(item, dict) for item in data):
        keys: List[str] = []
        seen = set()
        for item in data:
            for key in item.keys():
                if key in seen:
                    continue
                seen.add(key)
                keys.append(str(key))
        rows = [keys]
        rows.extend([[item.get(key) for key in keys] for item in data])
        table = _rows_to_markdown_table(rows, title="[JSON Records]")
        if table:
            return table

    return json.dumps(data, ensure_ascii=False, indent=2)


def _extract_text_from_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"pdf parser dependencies are not available: {exc}",
        ) from exc

    try:
        reader = PdfReader(path)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open pdf: {exc}") from exc

    page_texts: List[str] = []
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text(extraction_mode="layout") or "").strip()
        except TypeError:
            try:
                text = (page.extract_text() or "").strip()
            except Exception:
                text = ""
        except Exception:
            text = ""
        if text:
            page_texts.append(f"[Page {page_index}]\n{text}")

    if not page_texts:
        raise AppError(
            ERR_CELERY_ERROR,
            "pdf text extraction produced empty content; scanned pdf OCR is not supported yet",
        )

    return "\n\n".join(page_texts)


def _iter_docx_table_texts(document) -> Iterable[str]:
    for table_index, table in enumerate(getattr(document, "tables", []), start=1):
        rows = []
        for row in table.rows:
            cells = []
            seen_cells = set()
            for cell in row.cells:
                cell_key = id(getattr(cell, "_tc", cell))
                if cell_key in seen_cells:
                    continue
                seen_cells.add(cell_key)
                cells.append(cell.text)
            rows.append(cells)

        table_text = _rows_to_markdown_table(rows, title="[Table {0}]".format(table_index))
        if table_text:
            yield table_text


def _extract_text_from_docx(path: str) -> str:
    try:
        from docx import Document
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"docx parser dependencies are not available: {exc}",
        ) from exc

    try:
        document = Document(path)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open docx: {exc}") from exc

    parts: List[str] = []
    for paragraph in document.paragraphs:
        text = (paragraph.text or "").strip()
        if text:
            parts.append(text)

    parts.extend(_iter_docx_table_texts(document))
    return "\n\n".join(parts)


def _extract_text_from_xlsx(path: str) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"xlsx parser dependencies are not available: {exc}",
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open xlsx: {exc}") from exc

    parts: List[str] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            table_text = _rows_to_markdown_table(
                rows,
                title="[Sheet: {0}]".format(worksheet.title),
            )
            if table_text:
                parts.append(table_text)
    finally:
        workbook.close()

    return "\n\n".join(parts)


def extract_text_from_document(path: str, filename: str) -> str:
    validate_supported_document_filename(filename)

    ext = get_document_extension(filename)
    if ext in PLAIN_TEXT_DOCUMENT_EXTENSIONS:
        return _extract_text_from_plain_file(path)
    if ext == "csv":
        return _extract_text_from_csv(path)
    if ext == "json":
        return _extract_text_from_json(path)
    if ext == "pdf":
        return _extract_text_from_pdf(path)
    if ext == "docx":
        return _extract_text_from_docx(path)
    if ext == "xlsx":
        return _extract_text_from_xlsx(path)

    raise AppError(
        ERR_CELERY_ERROR,
        "unsupported document format; currently supported: %s"
        % supported_document_extensions_text(),
    )
