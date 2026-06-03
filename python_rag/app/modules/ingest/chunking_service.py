import csv
import io
import json
import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Sequence

from python_rag.app.core.error_codes import ERR_CELERY_ERROR
from python_rag.app.core.errors import AppError
from python_rag.app.shared.text_chunker import normalize_text, simple_chunk_text


SUPPORTED_DOCUMENT_EXTENSIONS = (
    "md",
    "txt",
    "json",
    "csv",
    "pdf",
    "docx",
    "xlsx",
)

PLAIN_TEXT_DOCUMENT_EXTENSIONS = {"md", "txt"}
TITLE_CHUNK_DOCUMENT_EXTENSIONS = {"md", "docx", "pdf"}

_MARKDOWN_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*#*\s*$")
_SETEXT_HEADING_RE = re.compile(r"^\s*(=+|-+)\s*$")
_NUMBERED_HEADING_RE = re.compile(
    r"^((?:\d+(?:\.\d+)*[.)]?|[A-Z][.)]))\s+(.+)$"
)
_CJK_HEADING_RE = re.compile(r"^(第[一二三四五六七八九十百千万\d]+[章节篇部分]|[一二三四五六七八九十]+[、.])\s*(.+)$")
_FENCED_CODE_RE = re.compile(r"^\s*(```|~~~)")
_FENCED_CODE_BLOCK_RE = re.compile(
    r"(^|\n)\s*(```|~~~).*?\n.*?\n\s*\2\s*(?=\n|$)",
    re.DOTALL,
)


@dataclass
class Chunk:
    content: str = ""
    metadata: Dict[str, str] = field(default_factory=dict)


def get_document_extension(filename: str) -> str:
    _, ext = os.path.splitext(filename or "")
    return ext.lower().lstrip(".")


def supported_document_extensions_text() -> str:
    return ", ".join(f".{ext}" for ext in SUPPORTED_DOCUMENT_EXTENSIONS)


def is_supported_document_filename(filename: str) -> bool:
    return get_document_extension(filename) in SUPPORTED_DOCUMENT_EXTENSIONS


def validate_supported_document_filename(filename: str) -> None:
    if is_supported_document_filename(filename):
        return
    raise AppError(
        ERR_CELERY_ERROR,
        "unsupported document format; currently supported: %s"
        % supported_document_extensions_text(),
    )


def _read_binary_file(path: str) -> bytes:
    if not os.path.exists(path):
        raise AppError(ERR_CELERY_ERROR, "document file does not exist")

    with open(path, "rb") as f:
        return f.read()


def _decode_text_bytes(raw: bytes) -> str:
    if not raw:
        return ""

    for encoding in ("utf-8", "utf-8-sig", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except Exception:
            continue

    return raw.decode("utf-8", errors="ignore")


def _extract_text_from_plain_file(path: str) -> str:
    return _decode_text_bytes(_read_binary_file(path))


def _compact_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(line.strip() for line in text.split("\n") if line.strip()).strip()


def _escape_markdown_table_cell(value: Any) -> str:
    return _compact_cell_text(value).replace("|", "\\|")


def _normalize_table_rows(rows: Iterable[Sequence[Any]]) -> List[List[str]]:
    normalized_rows: List[List[str]] = []
    max_width = 0

    for row in rows:
        cells = [_escape_markdown_table_cell(cell) for cell in row]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        normalized_rows.append(cells)
        max_width = max(max_width, len(cells))

    if max_width <= 0:
        return []

    for row in normalized_rows:
        if len(row) < max_width:
            row.extend([""] * (max_width - len(row)))
    return normalized_rows


def _rows_to_markdown_table(rows: Iterable[Sequence[Any]], title: str | None = None) -> str:
    normalized_rows = _normalize_table_rows(rows)
    if not normalized_rows:
        return ""

    header = normalized_rows[0]
    header = [
        cell if cell else "Column {0}".format(index + 1)
        for index, cell in enumerate(header)
    ]
    body_rows = normalized_rows[1:]

    parts: List[str] = []
    if title:
        parts.append(title)
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join(["---"] * len(header)) + " |")
    for row in body_rows:
        parts.append("| " + " | ".join(row) + " |")

    return "\n".join(parts)


def _extract_text_from_csv(path: str) -> str:
    raw_text = _decode_text_bytes(_read_binary_file(path))
    if not raw_text.strip():
        return ""

    sample = raw_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel

    try:
        reader = csv.reader(io.StringIO(raw_text), dialect)
        table = _rows_to_markdown_table(reader, title="[CSV Table]")
    except Exception:
        table = ""
    return table or raw_text


def _extract_text_from_json(path: str) -> str:
    raw_text = _decode_text_bytes(_read_binary_file(path))
    if not raw_text.strip():
        return ""

    try:
        data = json.loads(raw_text)
    except Exception:
        return raw_text

    if isinstance(data, list) and data and all(isinstance(item, dict) for item in data):
        keys: List[str] = []
        seen = set()
        for item in data:
            for key in item.keys():
                if key in seen:
                    continue
                seen.add(key)
                keys.append(str(key))
        rows = [keys]
        rows.extend([[item.get(key) for key in keys] for item in data])
        table = _rows_to_markdown_table(rows, title="[JSON Records]")
        if table:
            return table

    return json.dumps(data, ensure_ascii=False, indent=2)


def _extract_text_from_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"pdf parser dependencies are not available: {exc}",
        ) from exc

    try:
        reader = PdfReader(path)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open pdf: {exc}") from exc

    page_texts: List[str] = []
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text(extraction_mode="layout") or "").strip()
        except TypeError:
            try:
                text = (page.extract_text() or "").strip()
            except Exception:
                text = ""
        except Exception:
            text = ""
        if text:
            page_texts.append(f"<!-- Page {page_index} -->\n{text}")

    if not page_texts:
        raise AppError(
            ERR_CELERY_ERROR,
            "pdf text extraction produced empty content; scanned pdf OCR is not supported yet",
        )

    return "\n\n".join(page_texts)


def _iter_docx_table_texts(document) -> Iterable[str]:
    for table_index, table in enumerate(getattr(document, "tables", []), start=1):
        rows = []
        for row in table.rows:
            cells = []
            seen_cells = set()
            for cell in row.cells:
                cell_key = id(getattr(cell, "_tc", cell))
                if cell_key in seen_cells:
                    continue
                seen_cells.add(cell_key)
                cells.append(cell.text)
            rows.append(cells)

        table_text = _rows_to_markdown_table(rows, title="[Table {0}]".format(table_index))
        if table_text:
            yield table_text


def _docx_heading_level(paragraph) -> int | None:
    style = getattr(paragraph, "style", None)
    style_name = (getattr(style, "name", "") or "").strip().lower()
    style_id = (getattr(style, "style_id", "") or "").strip().lower()

    for value in (style_name, style_id):
        match = re.search(r"(?:heading|标题)\s*(\d+)", value)
        if match:
            return max(1, min(6, int(match.group(1))))

    return None


def _extract_text_from_docx(path: str) -> str:
    try:
        from docx import Document
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"docx parser dependencies are not available: {exc}",
        ) from exc

    try:
        document = Document(path)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open docx: {exc}") from exc

    parts: List[str] = []
    for paragraph in document.paragraphs:
        text = (paragraph.text or "").strip()
        if not text:
            continue

        heading_level = _docx_heading_level(paragraph)
        if heading_level:
            parts.append("{0} {1}".format("#" * heading_level, text))
        else:
            parts.append(text)

    parts.extend(_iter_docx_table_texts(document))
    return "\n\n".join(parts)


def _extract_text_from_xlsx(path: str) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise AppError(
            ERR_CELERY_ERROR,
            f"xlsx parser dependencies are not available: {exc}",
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise AppError(ERR_CELERY_ERROR, f"failed to open xlsx: {exc}") from exc

    parts: List[str] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            table_text = _rows_to_markdown_table(
                rows,
                title="[Sheet: {0}]".format(worksheet.title),
            )
            if table_text:
                parts.append(table_text)
    finally:
        workbook.close()

    return "\n\n".join(parts)


def extract_text_from_document(path: str, filename: str) -> str:
    validate_supported_document_filename(filename)

    ext = get_document_extension(filename)
    if ext in PLAIN_TEXT_DOCUMENT_EXTENSIONS:
        return _extract_text_from_plain_file(path)
    if ext == "csv":
        return _extract_text_from_csv(path)
    if ext == "json":
        return _extract_text_from_json(path)
    if ext == "pdf":
        return _extract_text_from_pdf(path)
    if ext == "docx":
        return _extract_text_from_docx(path)
    if ext == "xlsx":
        return _extract_text_from_xlsx(path)

    raise AppError(
        ERR_CELERY_ERROR,
        "unsupported document format; currently supported: %s"
        % supported_document_extensions_text(),
    )


def _is_page_marker(line: str) -> bool:
    stripped = line.strip()
    return bool(re.fullmatch(r"(?:<!-- Page \d+ -->|\[Page \d+\]|Page\s+\d+)", stripped))


def _has_section_content(lines: Sequence[str]) -> bool:
    return any(line.strip() and not _is_page_marker(line) for line in lines)


def _is_probable_heading(line: str, ext: str) -> tuple[int, str] | None:
    stripped = line.strip()
    if not stripped or _is_page_marker(stripped):
        return None

    markdown_match = _MARKDOWN_HEADING_RE.match(stripped)
    if markdown_match:
        title = markdown_match.group(2).strip()
        return len(markdown_match.group(1)), title

    if ext == "md":
        return None

    numbered_match = _NUMBERED_HEADING_RE.match(stripped)
    if numbered_match:
        marker = numbered_match.group(1)
        level = marker.count(".") + 1 if "." in marker else 1
        return max(1, min(6, level)), stripped

    cjk_match = _CJK_HEADING_RE.match(stripped)
    if cjk_match:
        marker = cjk_match.group(1)
        level = 1 if marker.startswith("第") else 2
        return level, stripped

    if len(stripped) <= 60 and stripped.endswith((":", "：")):
        return 3, stripped.rstrip(":：").strip()

    return None


def _render_heading(level: int, title: str) -> str:
    return "{0} {1}".format("#" * max(1, min(6, level)), title.strip())


def _is_fenced_code_marker(line: str) -> str | None:
    match = _FENCED_CODE_RE.match(line)
    if not match:
        return None
    return match.group(1)


def _length_excluding_code(text: str) -> int:
    total_length = 0
    last_end = 0
    for match in _FENCED_CODE_BLOCK_RE.finditer(text):
        start, end = match.span()
        total_length += len(text[last_end:start])
        last_end = end
    total_length += len(text[last_end:])
    return total_length


def _find_best_split_point(lines: Sequence[str]) -> int:
    if len(lines) <= 1:
        return -1

    for index in range(len(lines) - 2, 0, -1):
        if not lines[index].strip() and lines[index + 1].strip():
            if index > 0 and lines[index - 1].strip():
                return index + 1

    return len(lines) - 1


def _split_chunk_by_size(chunk: Chunk, chunk_size: int) -> List[Chunk]:
    if _length_excluding_code(chunk.content) <= chunk_size:
        return [chunk]

    sub_chunks: List[Chunk] = []
    current_lines: List[str] = []
    current_non_code_len = 0
    in_code = False
    code_fence: str | None = None

    for line in chunk.content.split("\n"):
        stripped = line.strip()
        fence = _is_fenced_code_marker(stripped)
        entering_code = bool(fence and not in_code)
        exiting_code = bool(fence and in_code and code_fence == fence)

        line_len = 0
        if not in_code and not entering_code:
            line_len = len(line) + 1
        elif exiting_code:
            line_len = len(line) + 1

        split_needed = (
            line_len > 0
            and current_non_code_len + line_len > chunk_size
            and bool(current_lines)
        )
        if split_needed:
            split_index = _find_best_split_point(current_lines)
            if split_index > 0:
                content = "\n".join(current_lines[:split_index]).strip()
                if content:
                    sub_chunks.append(Chunk(content=content, metadata=chunk.metadata.copy()))
                current_lines = current_lines[split_index:] + [line]
                current_non_code_len = _length_excluding_code("\n".join(current_lines))
            else:
                content = "\n".join(current_lines).strip()
                if content:
                    sub_chunks.append(Chunk(content=content, metadata=chunk.metadata.copy()))
                current_lines = [line]
                current_non_code_len = 0 if entering_code else line_len
        else:
            current_lines.append(line)
            current_non_code_len += line_len

        if entering_code:
            in_code = True
            code_fence = fence
        elif exiting_code:
            in_code = False
            code_fence = None

    if current_lines:
        content = "\n".join(current_lines).strip()
        if content:
            sub_chunks.append(Chunk(content=content, metadata=chunk.metadata.copy()))

    return sub_chunks or [chunk]


def _aggregate_chunks(chunks: Sequence[Chunk]) -> List[Chunk]:
    aggregated: List[Chunk] = []
    for chunk in chunks:
        content = chunk.content.strip()
        if not content:
            continue
        if aggregated and aggregated[-1].metadata == chunk.metadata:
            aggregated[-1].content = normalize_text(aggregated[-1].content + "\n" + content)
        else:
            aggregated.append(Chunk(content=content, metadata=chunk.metadata.copy()))
    return aggregated


def _is_heading_only_chunk(chunk: Chunk) -> bool:
    lines = [line.strip() for line in chunk.content.split("\n") if line.strip()]
    if not lines:
        return False
    return all(_MARKDOWN_HEADING_RE.match(line) for line in lines)


def _merge_heading_only_chunks(chunks: Sequence[Chunk]) -> List[Chunk]:
    merged: List[Chunk] = []
    pending_headings: List[str] = []
    pending_metadata: Dict[str, str] = {}

    for chunk in chunks:
        if _is_heading_only_chunk(chunk):
            pending_headings.extend(
                line.strip()
                for line in chunk.content.split("\n")
                if line.strip()
            )
            pending_metadata.update(chunk.metadata)
            continue

        if pending_headings:
            content = normalize_text(
                "\n".join(pending_headings) + "\n" + chunk.content
            )
            metadata = pending_metadata.copy()
            metadata.update(chunk.metadata)
            merged.append(Chunk(content=content, metadata=metadata))
            pending_headings = []
            pending_metadata = {}
        else:
            merged.append(chunk)

    return merged


def _split_into_heading_chunks(text: str, ext: str) -> List[Chunk]:
    text = normalize_text(text)
    if not text:
        return []

    lines = text.split("\n")
    chunks: List[Chunk] = []
    current_content: List[str] = []
    current_metadata: Dict[str, str] = {}
    header_stack: List[tuple[int, str, str]] = []
    seen_heading = False
    in_fenced_code = False
    code_fence: str | None = None
    index = 0

    def flush_current() -> None:
        nonlocal current_content
        if not current_content or not _has_section_content(current_content):
            current_content = []
            return
        chunks.append(
            Chunk(
                content=normalize_text("\n".join(current_content)),
                metadata=current_metadata.copy(),
            )
        )
        current_content = []

    while index < len(lines):
        line = lines[index]
        stripped = line.strip()
        fence = _is_fenced_code_marker(stripped)
        if fence and not in_fenced_code:
            in_fenced_code = True
            code_fence = fence
            heading = None
        elif fence and in_fenced_code and code_fence == fence:
            in_fenced_code = False
            code_fence = None
            heading = None
        elif in_fenced_code:
            heading = None
        else:
            heading = _is_probable_heading(line, ext)

        if ext == "md" and not in_fenced_code and index + 1 < len(lines):
            next_line = lines[index + 1].strip()
            if line.strip() and _SETEXT_HEADING_RE.match(next_line):
                heading = (1 if next_line.startswith("=") else 2, line.strip())
                index += 1

        if heading:
            level, title = heading
            flush_current()

            while header_stack and header_stack[-1][0] >= level:
                header_stack.pop()
            header_name = "h{0}".format(level)
            header_stack.append((level, header_name, title))
            current_metadata = {name: data for _, name, data in header_stack}
            current_content = [_render_heading(level, title)]
            seen_heading = True
        else:
            if not _is_page_marker(line) and (line.strip() or current_content):
                current_content.append(line)

        index += 1

    flush_current()

    if not seen_heading:
        return []

    return _merge_heading_only_chunks(_aggregate_chunks(chunks))


def chunk_text_by_title(
    text: str,
    filename: str,
    chunk_size: int = 800,
    overlap: int = 100,
) -> List[str]:
    ext = get_document_extension(filename)
    text = normalize_text(text)
    if not text:
        return []

    if chunk_size <= 0:
        raise ValueError("chunk_size must be positive")
    if overlap < 0:
        raise ValueError("overlap must be >= 0")
    if overlap >= chunk_size:
        raise ValueError("overlap must be smaller than chunk_size")

    if ext not in TITLE_CHUNK_DOCUMENT_EXTENSIONS:
        return simple_chunk_text(text, chunk_size=chunk_size, overlap=overlap)

    title_chunks = _split_into_heading_chunks(text, ext)
    if not title_chunks:
        return simple_chunk_text(text, chunk_size=chunk_size, overlap=overlap)

    chunks: List[str] = []
    for chunk in title_chunks:
        chunks.extend(
            sub_chunk.content
            for sub_chunk in _split_chunk_by_size(chunk, chunk_size=chunk_size)
        )

    return [chunk for chunk in chunks if chunk.strip()]
